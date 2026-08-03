from __future__ import annotations

import io
import subprocess
import tempfile
import zipfile
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from app.services.text_extraction import (
    ExtractedPage,
    PdfExtractionError,
    PdfExtractionResult,
    extract_pdf_text,
)
from app.services.tencent_ocr_provider import (
    TencentOcrConfig,
    TencentOcrError,
    extract_image_text,
)


class DocumentExtractionError(RuntimeError):
    """A stable, UI-safe reason why a document could not be normalized."""

    def __init__(
        self,
        error_code: str,
        *,
        source_page_count: int = 0,
        ocr_attempted_page_count: int = 0,
        ocr_successful_page_count: int = 0,
        ocr_selected_page_count: int = 0,
        ocr_failed_page_count: int = 0,
    ) -> None:
        super().__init__(error_code)
        # Keep only count-level observability through an error path.  The
        # worker can record this without retaining a document-level trace.
        self.source_page_count = source_page_count
        self.ocr_attempted_page_count = ocr_attempted_page_count
        self.ocr_successful_page_count = ocr_successful_page_count
        self.ocr_selected_page_count = ocr_selected_page_count
        self.ocr_failed_page_count = ocr_failed_page_count


def _document_error_from_pdf_error(exc: PdfExtractionError) -> DocumentExtractionError:
    """Preserve count-only OCR telemetry when a PDF result is rejected late."""

    return DocumentExtractionError(
        str(exc),
        source_page_count=exc.source_page_count,
        ocr_attempted_page_count=exc.ocr_attempted_page_count,
        ocr_successful_page_count=exc.ocr_successful_page_count,
        ocr_selected_page_count=exc.ocr_selected_page_count,
        ocr_failed_page_count=exc.ocr_failed_page_count,
    )


SUPPORTED_DOCUMENT_EXTENSIONS = frozenset(
    {
        ".pdf",
        ".doc",
        ".docx",
        ".png",
        ".jpg",
        ".jpeg",
        ".xls",
        ".xlsx",
        ".html",
        ".htm",
    }
)

_PDF_MAGIC = b"%PDF-"
_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"
_JPEG_MAGIC = b"\xff\xd8\xff"
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def validate_document_signature(*, filename: str, content: bytes) -> None:
    """Reject an extension/content mismatch before an original is persisted.

    This is intentionally a bounded preflight, not a parser: the expensive
    Office, OCR, PDF and spreadsheet work remains in the durable worker.  A
    DOC and XLS file share the OLE container signature, while DOCX/XLSX are
    verified as the expected Open XML package rather than merely a generic
    ZIP file.
    """

    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_DOCUMENT_EXTENSIONS:
        raise DocumentExtractionError("unsupported_document_type")
    if suffix == ".pdf":
        _require(content.startswith(_PDF_MAGIC))
        return
    if suffix == ".png":
        _require(content.startswith(_PNG_MAGIC))
        return
    if suffix in {".jpg", ".jpeg"}:
        _require(content.startswith(_JPEG_MAGIC))
        return
    if suffix in {".doc", ".xls"}:
        _require(content.startswith(_OLE_MAGIC))
        return
    if suffix == ".docx":
        _require_open_xml_package(content, required_member="word/document.xml")
        return
    if suffix == ".xlsx":
        _require_open_xml_package(content, required_member="xl/workbook.xml")
        return
    if suffix in {".html", ".htm"}:
        # HTML has no binary magic.  Requiring a leading markup token blocks
        # arbitrary binary data renamed to .html without excluding ordinary
        # HTML fragments such as `<section>…` or `<table>…`.
        prefix = content[:4096].decode("utf-8", errors="ignore").lstrip("\ufeff\x00\t\r\n ")
        _require(prefix.startswith("<"))
        return
    raise DocumentExtractionError("unsupported_document_type")


def validate_document_path_signature(*, path: Path, filename: str | None = None) -> None:
    """Repeat a bounded signature gate immediately before worker parsing.

    Upload handlers already validate bytes before an original is persisted.
    The worker repeats this inexpensive check because it is the final boundary
    before a legacy upload or a parser-repair copy reaches an untrusted file
    parser.  It reads only a small prefix except for Open XML's central
    directory, which ``ZipFile`` can inspect directly from the bounded file.
    """

    suffix = Path(filename or path.name).suffix.lower()
    if suffix not in SUPPORTED_DOCUMENT_EXTENSIONS:
        raise DocumentExtractionError("unsupported_document_type")
    try:
        with path.open("rb") as source:
            prefix = source.read(4096)
    except OSError as exc:
        raise DocumentExtractionError("document_open_failed") from exc
    if suffix == ".pdf":
        _require(prefix.startswith(_PDF_MAGIC))
        return
    if suffix == ".png":
        _require(prefix.startswith(_PNG_MAGIC))
        return
    if suffix in {".jpg", ".jpeg"}:
        _require(prefix.startswith(_JPEG_MAGIC))
        return
    if suffix in {".doc", ".xls"}:
        _require(prefix.startswith(_OLE_MAGIC))
        return
    if suffix == ".docx":
        _require_open_xml_path(path, required_member="word/document.xml")
        return
    if suffix == ".xlsx":
        _require_open_xml_path(path, required_member="xl/workbook.xml")
        return
    if suffix in {".html", ".htm"}:
        normalized = prefix.decode("utf-8", errors="ignore").lstrip(
            "\ufeff\x00\t\r\n "
        )
        _require(normalized.startswith("<"))
        return
    raise DocumentExtractionError("unsupported_document_type")


def _require(condition: bool) -> None:
    if not condition:
        raise DocumentExtractionError("invalid_document_signature")


def _require_open_xml_package(content: bytes, *, required_member: str) -> None:
    if not content.startswith(b"PK\x03\x04"):
        raise DocumentExtractionError("invalid_document_signature")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            _require_open_xml_members(archive, required_member=required_member)
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise DocumentExtractionError("invalid_document_signature") from exc


def _require_open_xml_path(path: Path, *, required_member: str) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            _require_open_xml_members(archive, required_member=required_member)
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise DocumentExtractionError("invalid_document_signature") from exc


def _require_open_xml_members(
    archive: zipfile.ZipFile,
    *,
    required_member: str,
) -> None:
    # A 15 MiB upload could otherwise contain an excessive central directory
    # even before its decompression budget is checked by the extractor.
    members = archive.infolist()
    if len(members) > 10_000:
        raise DocumentExtractionError("invalid_document_signature")
    member_names = {member.filename for member in members}
    _require("[Content_Types].xml" in member_names and required_member in member_names)


def extract_document_text(
    path: Path,
    *,
    min_text_chars_per_page: int,
    ocr_sparse_text_chars_per_page: int,
    tencent_ocr_config: TencentOcrConfig | None = None,
    max_pages: int = 30,
    max_text_chars: int = 250_000,
    max_archive_uncompressed_bytes: int = 100 * 1024 * 1024,
    max_spreadsheet_sheets: int = 20,
    max_spreadsheet_rows_per_sheet: int = 5_000,
    max_spreadsheet_cells: int = 50_000,
    office_timeout_seconds: int = 90,
) -> PdfExtractionResult:
    """Normalize one supported original into source-cited page text.

    Callers must run this only from the document extraction worker.  Every
    high-cost or attacker-controlled parser receives a deterministic budget;
    raising a `DocumentExtractionError` is preferable to silently truncating
    evidence and later making an AI conclusion from incomplete source text.
    """

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        try:
            return extract_pdf_text(
                path,
                min_text_chars_per_page=min_text_chars_per_page,
                ocr_sparse_text_chars_per_page=ocr_sparse_text_chars_per_page,
                tencent_ocr_config=tencent_ocr_config,
                max_pages=max_pages,
                max_text_chars=max_text_chars,
            )
        except PdfExtractionError as exc:
            raise _document_error_from_pdf_error(exc) from exc
    if suffix in {".doc", ".docx"}:
        if suffix == ".docx":
            _assert_zip_budget(
                path,
                max_uncompressed_bytes=max_archive_uncompressed_bytes,
                error_code="office_archive_limit_exceeded",
            )
        return _extract_office_as_pdf(
            path,
            minimum=min_text_chars_per_page,
            sparse=ocr_sparse_text_chars_per_page,
            config=tencent_ocr_config,
            max_pages=max_pages,
            max_text_chars=max_text_chars,
            timeout_seconds=office_timeout_seconds,
        )
    if suffix in {".xls", ".xlsx"}:
        if suffix == ".xls":
            return _extract_legacy_spreadsheet(
                path,
                minimum=min_text_chars_per_page,
                max_text_chars=max_text_chars,
                max_archive_uncompressed_bytes=max_archive_uncompressed_bytes,
                max_spreadsheet_sheets=max_spreadsheet_sheets,
                max_spreadsheet_rows_per_sheet=max_spreadsheet_rows_per_sheet,
                max_spreadsheet_cells=max_spreadsheet_cells,
                timeout_seconds=office_timeout_seconds,
            )
        return _extract_spreadsheet(
            path,
            minimum=min_text_chars_per_page,
            max_text_chars=max_text_chars,
            max_archive_uncompressed_bytes=max_archive_uncompressed_bytes,
            max_spreadsheet_sheets=max_spreadsheet_sheets,
            max_spreadsheet_rows_per_sheet=max_spreadsheet_rows_per_sheet,
            max_spreadsheet_cells=max_spreadsheet_cells,
        )
    if suffix in {".html", ".htm"}:
        return _extract_html(path, minimum=min_text_chars_per_page, max_text_chars=max_text_chars)
    if suffix in {".png", ".jpg", ".jpeg"}:
        return _extract_image(
            path,
            minimum=min_text_chars_per_page,
            max_text_chars=max_text_chars,
            config=tencent_ocr_config,
        )
    raise DocumentExtractionError("unsupported_document_type")


def _result(
    texts: list[str],
    *,
    minimum: int,
    parser: str,
    max_pages: int,
    max_text_chars: int,
    ocr_attempted_page_count: int = 0,
    ocr_successful_page_count: int = 0,
    ocr_selected_page_count: int = 0,
    ocr_failed_page_count: int = 0,
) -> PdfExtractionResult:
    if len(texts) > max_pages:
        raise DocumentExtractionError("document_page_limit_exceeded")
    normalized_texts = [text.strip() for text in texts]
    total_text_chars = sum(len(text) for text in normalized_texts)
    if total_text_chars > max_text_chars:
        raise DocumentExtractionError("document_text_limit_exceeded")
    pages = [
        ExtractedPage(
            page_no=index,
            text=text,
            non_whitespace_chars=len("".join(text.split())),
        )
        for index, text in enumerate(normalized_texts, 1)
    ]
    flags = [
        f"page_{page.page_no}_insufficient_text"
        for page in pages
        if page.non_whitespace_chars < minimum
    ]
    if not any(page.text for page in pages):
        flags.append("no_extractable_text")
    return PdfExtractionResult(
        source_page_count=len(pages),
        parsed_page_count=sum(page.non_whitespace_chars >= minimum for page in pages),
        pages=pages,
        raw_text="\n\n".join(
            f"--- PAGE {page.page_no} ---\n{page.text}"
            for page in pages
            if page.text
        ),
        quality_flags=flags,
        parser_version=parser,
        ocr_attempted_page_count=ocr_attempted_page_count,
        ocr_successful_page_count=ocr_successful_page_count,
        ocr_selected_page_count=ocr_selected_page_count,
        ocr_failed_page_count=ocr_failed_page_count,
    )


def _extract_office_as_pdf(
    path: Path,
    *,
    minimum: int,
    sparse: int,
    config: TencentOcrConfig | None,
    max_pages: int,
    max_text_chars: int,
    timeout_seconds: int,
) -> PdfExtractionResult:
    with tempfile.TemporaryDirectory() as directory:
        output = Path(directory)
        try:
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(output),
                    str(path),
                ],
                check=True,
                capture_output=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise DocumentExtractionError("office_conversion_timed_out") from exc
        except (OSError, subprocess.SubprocessError) as exc:
            raise DocumentExtractionError("office_conversion_failed") from exc
        converted = output / f"{path.stem}.pdf"
        if not converted.is_file():
            raise DocumentExtractionError("office_conversion_failed")
        try:
            return extract_pdf_text(
                converted,
                min_text_chars_per_page=minimum,
                ocr_sparse_text_chars_per_page=sparse,
                tencent_ocr_config=config,
                max_pages=max_pages,
                max_text_chars=max_text_chars,
            )
        except PdfExtractionError as exc:
            raise _document_error_from_pdf_error(exc) from exc


def _assert_zip_budget(
    path: Path,
    *,
    max_uncompressed_bytes: int,
    error_code: str,
) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            total = 0
            for member in archive.infolist():
                total += member.file_size
                if total > max_uncompressed_bytes:
                    raise DocumentExtractionError(error_code)
    except DocumentExtractionError:
        raise
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise DocumentExtractionError("spreadsheet_open_failed") from exc


def _extract_spreadsheet(
    path: Path,
    *,
    minimum: int,
    max_text_chars: int,
    max_archive_uncompressed_bytes: int,
    max_spreadsheet_sheets: int,
    max_spreadsheet_rows_per_sheet: int,
    max_spreadsheet_cells: int,
) -> PdfExtractionResult:
    _assert_zip_budget(
        path,
        max_uncompressed_bytes=max_archive_uncompressed_bytes,
        error_code="spreadsheet_archive_limit_exceeded",
    )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentExtractionError("spreadsheet_open_failed") from exc
    try:
        if len(workbook.worksheets) > max_spreadsheet_sheets:
            raise DocumentExtractionError("spreadsheet_sheet_limit_exceeded")
        texts: list[str] = []
        total_cells = 0
        total_text_chars = 0
        for sheet in workbook.worksheets:
            if sheet.max_row > max_spreadsheet_rows_per_sheet:
                raise DocumentExtractionError("spreadsheet_row_limit_exceeded")
            projected_cells = sheet.max_row * sheet.max_column
            if projected_cells > max_spreadsheet_cells - total_cells:
                raise DocumentExtractionError("spreadsheet_cell_limit_exceeded")
            page_lines: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                total_cells += len(row)
                if total_cells > max_spreadsheet_cells:
                    raise DocumentExtractionError("spreadsheet_cell_limit_exceeded")
                values = [str(value).strip() for value in row if value not in (None, "")]
                if not values:
                    continue
                line = " | ".join(values)
                total_text_chars += len(line) + 1
                if total_text_chars > max_text_chars:
                    raise DocumentExtractionError("document_text_limit_exceeded")
                page_lines.append(line)
            texts.append("\n".join(page_lines))
    finally:
        workbook.close()
    return _result(
        texts or [""],
        minimum=minimum,
        parser="openpyxl",
        max_pages=max_spreadsheet_sheets,
        max_text_chars=max_text_chars,
    )


def _extract_legacy_spreadsheet(
    path: Path,
    *,
    minimum: int,
    max_text_chars: int,
    max_archive_uncompressed_bytes: int,
    max_spreadsheet_sheets: int,
    max_spreadsheet_rows_per_sheet: int,
    max_spreadsheet_cells: int,
    timeout_seconds: int,
) -> PdfExtractionResult:
    """Convert legacy XLS in the worker, then apply XLSX quotas."""

    with tempfile.TemporaryDirectory() as directory:
        output = Path(directory)
        try:
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(output),
                    str(path),
                ],
                check=True,
                capture_output=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise DocumentExtractionError("spreadsheet_conversion_timed_out") from exc
        except (OSError, subprocess.SubprocessError) as exc:
            raise DocumentExtractionError("spreadsheet_conversion_failed") from exc
        converted = output / f"{path.stem}.xlsx"
        if not converted.is_file():
            raise DocumentExtractionError("spreadsheet_conversion_failed")
        return _extract_spreadsheet(
            converted,
            minimum=minimum,
            max_text_chars=max_text_chars,
            max_archive_uncompressed_bytes=max_archive_uncompressed_bytes,
            max_spreadsheet_sheets=max_spreadsheet_sheets,
            max_spreadsheet_rows_per_sheet=max_spreadsheet_rows_per_sheet,
            max_spreadsheet_cells=max_spreadsheet_cells,
        )


def _extract_html(path: Path, *, minimum: int, max_text_chars: int) -> PdfExtractionResult:
    try:
        soup = BeautifulSoup(path.read_bytes(), "html.parser")
    except OSError as exc:
        raise DocumentExtractionError("html_open_failed") from exc
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return _result(
        [soup.get_text("\n", strip=True)],
        minimum=minimum,
        parser="beautifulsoup4",
        max_pages=1,
        max_text_chars=max_text_chars,
    )


def _extract_image(
    path: Path,
    *,
    minimum: int,
    max_text_chars: int,
    config: TencentOcrConfig | None,
) -> PdfExtractionResult:
    if config is None:
        raise DocumentExtractionError("tencent_ocr_not_configured")
    try:
        text = extract_image_text(path=path, config=config)
    except TencentOcrError as exc:
        raise DocumentExtractionError(str(exc)) from exc
    return _result(
        [text],
        minimum=minimum,
        parser="tencent-ocr",
        max_pages=1,
        max_text_chars=max_text_chars,
        ocr_attempted_page_count=1,
        ocr_successful_page_count=1,
        ocr_selected_page_count=1,
    )
